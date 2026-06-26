import os
import openpyxl
from copy import copy

def generate_water_template():
    src_path = os.path.join(os.path.dirname(__file__), "power_readings.xlsx")
    dst_path = os.path.join(os.path.dirname(__file__), "water_readings.xlsx")
    
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"Source power template not found at {src_path}")
        
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active # 'power_readings'
    ws.title = 'water_readings'
    
    # Save the original style of cell K1 (Doc Info) before unmerging
    k1_cell = ws.cell(row=1, column=11)
    k1_font = copy(k1_cell.font)
    k1_border = copy(k1_cell.border)
    k1_fill = copy(k1_cell.fill)
    k1_alignment = copy(k1_cell.alignment)
    k1_number_format = k1_cell.number_format

    # 1. Unmerge all existing merged cells to prevent issues when deleting columns/rows
    merged_ranges = list(ws.merged_cells.ranges)
    for r in merged_ranges:
        ws.unmerge_cells(str(r))
        
    # 2. Delete rows 36 onwards to remove the bottom table completely
    ws.delete_rows(36, 35)
    
    # 3. Discard the second image that was located in the deleted bottom table
    if hasattr(ws, '_images'):
        ws._images = [img for img in ws._images if getattr(img.anchor._from, 'row', 0) < 35]
    
    # 4. Copy cell styles from column B (2) to new columns N, O, P, Q (columns 14, 15, 16, 17)
    # for all remaining rows (1 to 35)
    for r in range(1, 36):
        source_cell = ws.cell(row=r, column=2)
        for col_idx in range(14, 18):
            target_cell = ws.cell(row=r, column=col_idx)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                
    # 5. Set column widths for new columns N, O, P, Q to match column B
    width_b = ws.column_dimensions['B'].width if ws.column_dimensions['B'].width else 12
    for col_letter in ['N', 'O', 'P', 'Q']:
        ws.column_dimensions[col_letter].width = width_b
        
    # 6. Write new header values
    # Title and Doc Info (Row 1)
    ws.cell(row=1, column=1, value="               BARANI HYDRAULICS INDIA PRIVATE LIMITED")
    
    # Set Doc Info cell at P1 with the copied K1 style and wrap_text=True
    p1_cell = ws.cell(row=1, column=16, value="DOC NO: R/MAI/EB\nMONTH/YEAR: ") # Column P (16)
    p1_cell.font = k1_font
    p1_cell.border = k1_border
    p1_cell.fill = k1_fill
    from openpyxl.styles import Alignment
    p1_cell.alignment = Alignment(
        horizontal=k1_alignment.horizontal if k1_alignment else 'left',
        vertical=k1_alignment.vertical if k1_alignment else 'center',
        wrap_text=True
    )
    p1_cell.number_format = k1_number_format

    
    # Subheader (Row 2 and Row 19)
    ws.cell(row=2, column=2, value="WATER VALVE READINGS")
    ws.cell(row=19, column=2, value="WATER VALVE READINGS")
    
    # Column headers (Row 3 and Row 20)
    ws.cell(row=3, column=1, value="S.NO")
    ws.cell(row=20, column=1, value="S.NO")
    for i in range(1, 17):
        ws.cell(row=3, column=i+1, value=f"VALVE {i}")
        ws.cell(row=20, column=i+1, value=f"VALVE {i}")
        
    # 7. Re-merge new ranges
    new_merges = [
        "A1:O1", "P1:Q1",
        "B2:Q2", "B19:Q19"
    ]
    for r in new_merges:
        ws.merge_cells(r)
        
    # 8. Clear data cells in rows 4-18 and 21-35 for columns B to Q (2 to 17)
    for r in range(4, 19):
        for c in range(2, 18):
            ws.cell(row=r, column=c).value = None
    for r in range(21, 36):
        for c in range(2, 18):
            ws.cell(row=r, column=c).value = None
            
    wb.save(dst_path)
    print("Successfully generated water_readings.xlsx template!")

if __name__ == "__main__":
    generate_water_template()
