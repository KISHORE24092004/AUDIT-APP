import os
import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter

def generate_utility_template(utility_name, doc_no, sheet_title, headers, col_count):
    src_path = os.path.join(os.path.dirname(__file__), "power_readings.xlsx")
    dst_name = f"{utility_name}_readings_log.xlsx" if "waste" not in utility_name else f"{utility_name}_log.xlsx"
    dst_path = os.path.join(os.path.dirname(__file__), dst_name)
    
    if not os.path.exists(src_path):
        raise FileNotFoundError(f"Source power template not found at {src_path}")
        
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active
    ws.title = f"{utility_name}_readings" if "waste" not in utility_name else f"{utility_name}"
    
    # Save the original style of cell K1 (Doc Info) before unmerging
    k1_cell = ws.cell(row=1, column=11)
    k1_font = copy(k1_cell.font)
    k1_border = copy(k1_cell.border)
    k1_fill = copy(k1_cell.fill)
    k1_alignment = copy(k1_cell.alignment)
    k1_number_format = k1_cell.number_format

    # 1. Unmerge all existing merged cells
    merged_ranges = list(ws.merged_cells.ranges)
    for r in merged_ranges:
        ws.unmerge_cells(str(r))
        
    # 2. Delete rows 36 onwards
    ws.delete_rows(36, 35)
    
    # 3. Discard the second image
    if hasattr(ws, '_images'):
        ws._images = [img for img in ws._images if getattr(img.anchor._from, 'row', 0) < 35]
    
    # 4. Copy cell styles from column B (2) to target columns (up to col_count)
    for r in range(1, 36):
        source_cell = ws.cell(row=r, column=2)
        for col_idx in range(12, col_count + 1):
            target_cell = ws.cell(row=r, column=col_idx)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                
    # 5. Set column widths for target columns
    width_b = ws.column_dimensions['B'].width if ws.column_dimensions['B'].width else 12
    for col_idx in range(12, col_count + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width_b
        
    # 6. Write new header values
    ws.cell(row=1, column=1, value="               BARANI HYDRAULICS INDIA PRIVATE LIMITED")
    
    # Set Doc Info cell at the last column with the copied K1 style and wrap_text=True
    doc_cell = ws.cell(row=1, column=col_count, value=f"DOC NO: {doc_no}\nMONTH/YEAR: ")
    doc_cell.font = k1_font
    doc_cell.border = k1_border
    doc_cell.fill = k1_fill
    from openpyxl.styles import Alignment
    doc_cell.alignment = Alignment(
        horizontal=k1_alignment.horizontal if k1_alignment else 'left',
        vertical=k1_alignment.vertical if k1_alignment else 'center',
        wrap_text=True
    )
    doc_cell.number_format = k1_number_format

    # Subheader (Row 2 and Row 19)
    ws.cell(row=2, column=2, value=sheet_title.upper())
    ws.cell(row=19, column=2, value=sheet_title.upper())
    
    # Column headers (Row 3 and Row 20)
    ws.cell(row=3, column=1, value="S.NO")
    ws.cell(row=20, column=1, value="S.NO")
    
    for idx, header in enumerate(headers):
        col_idx = idx + 2
        ws.cell(row=3, column=col_idx, value=header)
        ws.cell(row=20, column=col_idx, value=header)
        
    # 7. Re-merge ranges
    last_col_letter = get_column_letter(col_count)
    if col_count > 1:
        prev_col_letter = get_column_letter(col_count - 1)
        new_merges = [
            f"A1:{prev_col_letter}1", f"{last_col_letter}1:{last_col_letter}1",
            f"B2:{last_col_letter}2", f"B19:{last_col_letter}19"
        ]
    else:
        new_merges = [
            f"A1:{last_col_letter}1",
            f"B2:{last_col_letter}2", f"B19:{last_col_letter}19"
        ]
        
    for r in new_merges:
        ws.merge_cells(r)
        
    # 8. Clear data cells in rows 4-18 and 21-35 for columns B to last_col
    for r in range(4, 19):
        for c in range(2, col_count + 1):
            ws.cell(row=r, column=c).value = None
    for r in range(21, 36):
        for c in range(2, col_count + 1):
            ws.cell(row=r, column=c).value = None
            
    # 9. Clean/Delete any columns beyond col_count in the workbook
    if ws.max_column > col_count:
        ws.delete_cols(col_count + 1, ws.max_column - col_count)

    wb.save(dst_path)
    print(f"Successfully generated {dst_name} template!")

if __name__ == "__main__":
    genset_headers = [
        "BATTERY VOLTAGE", "DIESEL FILLING (LTR)", "RUNNING HOURS", "VOLTAGE",
        "KW/H", "DIESEL LEVEL", "RADIATOR WATER", "CARETAKER SIGN"
    ]
    # Genset 125kW: Copy exact template
    import shutil
    shutil.copy(
        os.path.join(os.path.dirname(__file__), "125kW readings.xlsx"),
        os.path.join(os.path.dirname(__file__), "genset_125kw_readings_log.xlsx")
    )
    print("Successfully copied 125kW readings.xlsx as genset_125kw_readings_log.xlsx!")
    # Genset 160kW: Copy exact template
    shutil.copy(
        os.path.join(os.path.dirname(__file__), "160kW readings.xlsx"),
        os.path.join(os.path.dirname(__file__), "genset_160kw_readings_log.xlsx")
    )
    print("Successfully copied 160kW readings.xlsx as genset_160kw_readings_log.xlsx!")
    
    compressor_headers = [
        "RUNNING HOURS", "LOAD HOURS", "MOTOR HOURS", "BAR", "TEMPERATURE", "CARETAKER SIGN"
    ]
    # Compressor-1
    generate_utility_template("compressor1", "R/MAI/CR/01", "Compressor-1 Telemetry Readings Log", compressor_headers, 7)
    # Compressor-2
    generate_utility_template("compressor2", "R/MAI/CR/02", "Compressor-2 Telemetry Readings Log", compressor_headers, 7)
    
    # Canteen Waste: Copy exact template
    shutil.copy(
        os.path.join(os.path.dirname(__file__), "canteen_waste_template.xlsx"),
        os.path.join(os.path.dirname(__file__), "canteen_waste_log.xlsx")
    )
    print("Successfully copied canteen_waste_template.xlsx as canteen_waste_log.xlsx!")
